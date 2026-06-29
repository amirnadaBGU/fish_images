"""Build an .xlsx report from a measure.json ``meta`` dict.

The JSON is the source of truth; this regenerates the workbook each time, so it
is safe to call after the review UI edits keypoints. Columns, per fish/frame:
length, center of mass (x, y), top-left tip (x, y), bottom-right tip (x, y).
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def _columns(meta: dict):
    unit = meta.get("unit", "px")
    calibrated = bool(meta.get("px_per_unit"))
    cols = [("frame", "name"), ("frame #", "frame_index"), ("time (s)", "timestamp_s")]
    if calibrated:
        cols.append((f"length ({unit})", f"curved_length_{unit}"))
    cols += [
        ("curved length (px)", "curved_length_px"),
        ("straight length (px)", "straight_tip_to_tip_px"),
        ("bending ratio", "bending_ratio"),
        ("center of mass x", "com_x"),
        ("center of mass y", "com_y"),
        ("tip top-left x", "tip_tl_x"),
        ("tip top-left y", "tip_tl_y"),
        ("tip bottom-right x", "tip_br_x"),
        ("tip bottom-right y", "tip_br_y"),
        ("problematic", "problematic"),
        ("flags", "flags"),
        ("edited", "edited"),
    ]
    return cols


def write_report(meta: dict, xlsx_path) -> Path:
    xlsx_path = Path(xlsx_path)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "measurements"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    edited_fill = PatternFill("solid", fgColor="FFF2CC")
    problem_fill = PatternFill("solid", fgColor="F8CBAD")

    cols = _columns(meta)
    for c, (title, _) in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=c, value=title)
        cell.font = header_font
        cell.fill = header_fill
    ws.freeze_panes = "A2"

    for r, rec in enumerate(meta.get("frames", []), start=2):
        for c, (_, key) in enumerate(cols, start=1):
            ws.cell(row=r, column=c, value=rec.get(key))
        # problematic (orange) takes priority; edited (yellow) otherwise
        if rec.get("problematic"):
            for c in range(1, len(cols) + 1):
                ws.cell(row=r, column=c).fill = problem_fill
        elif rec.get("edited"):
            for c in range(1, len(cols) + 1):
                ws.cell(row=r, column=c).fill = edited_fill

    for c, (title, _) in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(c)].width = max(12, len(title) + 2)

    # Summary sheet
    s = wb.create_sheet("summary")
    # summary stats use only good frames (skip problematic + NaN)
    lengths = [v for rec in meta.get("frames", []) if not rec.get("problematic")
               for v in [rec.get("curved_length_px")]
               if isinstance(v, (int, float)) and v == v]  # v==v drops NaN
    n_problem = sum(1 for rec in meta.get("frames", []) if rec.get("problematic"))
    rows = [
        ("video", meta.get("video")),
        ("frames measured", len(meta.get("frames", []))),
        ("problematic frames", n_problem),
        ("sampling (every N)", meta.get("every")),
        ("keypoints", meta.get("keypoints")),
        ("px per unit", meta.get("px_per_unit")),
        ("unit", meta.get("unit")),
        ("mean curved length (px)", round(sum(lengths) / len(lengths), 1) if lengths else None),
        ("min curved length (px)", round(min(lengths), 1) if lengths else None),
        ("max curved length (px)", round(max(lengths), 1) if lengths else None),
    ]
    for r, (kk, vv) in enumerate(rows, start=1):
        s.cell(row=r, column=1, value=kk).font = Font(bold=True)
        s.cell(row=r, column=2, value=vv)
    s.column_dimensions["A"].width = 24
    s.column_dimensions["B"].width = 28

    wb.save(xlsx_path)
    return xlsx_path
